"""
export_excel.py  –  Semua logika export ke Excel

STRUKTUR SHEET SIMPANAN (diperbaiki):
  - "Rekap Simpanan"     : ringkasan per anggota, semua jenis
  - "Detail [Jenis]"     : list transaksi per jenis dgn kolom Tgl Bayar (PERBAIKAN 1)
  - "Matriks [Jenis]"    : matriks Jan-Des per tahun
  - "Rekap Tahunan"      : rekap per periode semua jenis (PERBAIKAN 3 – tidak ada sheet per-periode)
  - "Smp-Semua"          : akumulatif total semua periode

LOGIKA REKAP TAHUNAN (PERBAIKAN 3):
  - pokok, wajib, sukarela  → AKUMULATIF (semua periode dijumlah terus)
  - hariraya, khusus        → PER TAHUN (hanya tahun yang sama, bukan per periode)

PERBAIKAN 2: Angsuran tampilkan Bunga (Rp) dan Total Bayar (Pokok+Bunga)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from helpers import JENIS_LABEL, JENIS_LIST, BULAN_NAMA, BULAN_SHORT, hitung_angsuran

# ── Palet warna ───────────────────────────────────────────────────────────────
HEX = dict(
    navy="1E3A5F", blue="2B6CB0", blue_lt="EBF8FF", blue_mid="DBEAFE",
    green="276749", green_lt="F0FFF4",
    orange="C05621", orange_lt="FFFAF0",
    teal="0D6EAE", teal_lt="E0F2FE",
    purple="553C9A", purple_lt="F3E8FF",
    amber="92400E", amber_lt="FFFBEB",
    red="991B1B", red_lt="FEF2F2",
    gray="374151", gray_lt="F9FAFB",
    total_bg="FED7AA", total_fg="7B341E",
    white="FFFFFF", alt="F0F7FF",
    border="B0C4DE",
)
FMT_RP   = '#,##0'
FMT_RP_Z = '#,##0;(#,##0);"-"'

def _f(bold=False, color="000000", size=10, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)
def _fill(h): return PatternFill("solid", fgColor=h)
def _al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

thin = Side(style="thin",   color=HEX["border"])
med  = Side(style="medium", color=HEX["navy"])
def _bord(): return Border(top=thin, bottom=thin, left=thin, right=thin)
def _bord_h(): return Border(top=med, bottom=med, left=thin, right=thin)

def _col(n): return get_column_letter(n)

def _title_block(ws, title, subtitle, info, ncols):
    for row, text, bg, fg, sz, bold in [
        (1, title,    HEX["navy"],    HEX["white"], 13, True),
        (2, subtitle, HEX["blue"],    HEX["white"], 10, True),
        (3, info,     HEX["blue_lt"], HEX["navy"],  9,  False),
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.font      = _f(bold=bold, color=fg, size=sz)
        c.fill      = _fill(bg)
        c.alignment = _al("center" if row < 3 else "left")
        ws.row_dimensions[row].height = 22 if row == 1 else 18

def _header_row(ws, headers, row, col_colors=None):
    for i, h in enumerate(headers, 1):
        bg = col_colors[i-1] if col_colors and i-1 < len(col_colors) else HEX["blue"]
        c  = ws.cell(row=row, column=i, value=h)
        c.font      = _f(bold=True, color=HEX["white"], size=9)
        c.fill      = _fill(bg)
        c.alignment = _al("center")
        c.border    = _bord_h()
    ws.row_dimensions[row].height = 18

def _data_cells(ws, values, row, num_cols=None, alt=False):
    bg = HEX["alt"] if alt else HEX["white"]
    num_cols = num_cols or []
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.fill      = _fill(bg)
        c.border    = _bord()
        c.font      = _f(size=9)
        c.alignment = _al("right" if i in num_cols else "left")
        if i in num_cols and isinstance(v, (int, float)):
            c.number_format = FMT_RP_Z

def _total_cells(ws, values, row, num_cols=None):
    num_cols = num_cols or []
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font      = _f(bold=True, color=HEX["total_fg"], size=9)
        c.fill      = _fill(HEX["total_bg"])
        c.border    = _bord_h()
        c.alignment = _al("right" if i in num_cols else "center")
        if i in num_cols and isinstance(v, (int, float)):
            c.number_format = FMT_RP_Z
    ws.row_dimensions[row].height = 18

def _set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _freeze(ws, r, c):
    ws.freeze_panes = ws.cell(row=r, column=c)

def _no_grid(ws):
    ws.sheet_view.showGridLines = False

def _periode_info(conn, periode_id=None, tahun=None, bulan=None):
    parts = []
    if periode_id:
        r = conn.execute(
            "SELECT nama, tahun, bulan_mulai, bulan_akhir, tgl_mulai, tgl_akhir "
            "FROM periode WHERE id=?", (periode_id,)
        ).fetchone()
        if r:
            from calendar import monthrange
            from datetime import datetime as _dt
            # Gunakan tgl_mulai/tgl_akhir dari DB jika tersedia, fallback ke perhitungan bulan
            def _fmt(tgl_str, fallback_day, fallback_m, fallback_y):
                if tgl_str:
                    try:
                        d = _dt.strptime(tgl_str, "%Y-%m-%d")
                        return f"{d.day:02d} {BULAN_NAMA[d.month]} {d.year}"
                    except Exception:
                        pass
                return f"{fallback_day:02d} {BULAN_NAMA[fallback_m]} {fallback_y}"
            hari_akhir = monthrange(r['tahun'], r['bulan_akhir'])[1]
            tgl_mulai = _fmt(r['tgl_mulai'], 1, r['bulan_mulai'], r['tahun'])
            tgl_akhir = _fmt(r['tgl_akhir'], hari_akhir, r['bulan_akhir'], r['tahun'])
            parts.append(f"Periode: {r['nama']}  ({tgl_mulai} s/d {tgl_akhir})")
    if tahun:  parts.append(f"Tahun: {tahun}")
    if bulan:  parts.append(f"Bulan: {BULAN_NAMA[bulan]}")
    parts.append(f"Dicetak: {date.today().strftime('%d %B %Y')}")
    return "  |  ".join(parts)

def _where(periode_id=None, tahun=None, bulan=None, t="s"):
    conds, vals = [], []
    if periode_id: conds.append(f"{t}.periode_id=?"); vals.append(periode_id)
    if tahun:      conds.append(f"{t}.tahun=?");      vals.append(tahun)
    if bulan:      conds.append(f"{t}.bulan=?");      vals.append(bulan)
    return ("WHERE " + " AND ".join(conds)) if conds else "", vals

# ═══════════════════════════════════════════════════════════════════════
#  SHEET: ANGGOTA
# ═══════════════════════════════════════════════════════════════════════
def _sheet_anggota(wb, conn, periode_id=None):
    ws = wb.create_sheet("Anggota")
    _no_grid(ws)
    _title_block(ws, "KOPERASI LANGGENG", "DATA ANGGOTA",
                 _periode_info(conn, periode_id), 6)
    _header_row(ws, ["No","No. Anggota","Nama Lengkap","Alamat","No. HP","Tgl Bergabung"], 4)
    rows = conn.execute("SELECT * FROM anggota ORDER BY id").fetchall()
    for i, r in enumerate(rows, 1):
        _data_cells(ws, [i, r["no_anggota"], r["nama"],
                         r["alamat"] or "-", r["no_hp"] or "-", r["tgl_masuk"]], i+4, alt=i%2==0)
    _total_cells(ws, [f"Total: {len(rows)} anggota","","","","",""], len(rows)+5)
    _set_widths(ws, {"A":5,"B":12,"C":26,"D":28,"E":16,"F":14})


# ═══════════════════════════════════════════════════════════════════════
#  SHEET: REKAP SIMPANAN (ringkasan per anggota)
# ═══════════════════════════════════════════════════════════════════════
def _sheet_rekap_simpanan(wb, conn, periode_id=None, tahun=None, bulan=None):
    ws = wb.create_sheet("Rekap Simpanan")
    _no_grid(ws)
    _title_block(ws, "KOPERASI LANGGENG", "REKAP SIMPANAN PER ANGGOTA",
                 _periode_info(conn, periode_id, tahun, bulan), 10)
    _header_row(ws, ["No","No. Anggota","Nama",
                     "Pokok","Wajib","Sukarela","Khusus","Hari Raya","Total","Ket."], 4)
    anggota = conn.execute("SELECT * FROM anggota ORDER BY id").fetchall()
    totals  = {j: 0 for j in JENIS_LIST}
    wh, wv  = _where(periode_id, tahun, bulan)
    for i, a in enumerate(anggota, 1):
        vals = {}
        for j in JENIS_LIST:
            q = f"SELECT COALESCE(SUM(jumlah),0) FROM simpanan s WHERE s.anggota_id=? AND s.jenis=?"
            p = [a["id"], j]
            if wh: q += " AND " + wh[6:]; p += wv
            vals[j] = conn.execute(q, p).fetchone()[0] or 0
            totals[j] += vals[j]
        grand = sum(vals.values())
        has_pin = conn.execute(
            "SELECT COUNT(*) FROM pinjaman WHERE anggota_id=? AND status='aktif'",
            (a["id"],)).fetchone()[0]
        _data_cells(ws,
            [i, a["no_anggota"], a["nama"],
             vals["pokok"], vals["wajib"], vals["sukarela"],
             vals["khusus"], vals["hariraya"], grand,
             "Ada Pinjaman" if has_pin else ""],
            i+4, num_cols=[4,5,6,7,8,9], alt=i%2==0)
    nr = len(anggota)+5
    _total_cells(ws,
        ["","","TOTAL", totals["pokok"], totals["wajib"], totals["sukarela"],
         totals["khusus"], totals["hariraya"], sum(totals.values()), ""],
        nr, num_cols=[4,5,6,7,8,9])
    _set_widths(ws,{"A":4,"B":12,"C":24,"D":14,"E":14,"F":14,"G":14,"H":14,"I":16,"J":16})


# ═══════════════════════════════════════════════════════════════════════
#  SHEET: DETAIL SIMPANAN PER JENIS — dengan kolom TGL BAYAR
#  PERBAIKAN 1: tgl diambil langsung dari field tgl (NOT NULL di DB)
# ═══════════════════════════════════════════════════════════════════════
def _sheet_detail_simpanan(wb, conn, jenis, periode_id=None, tahun=None, bulan=None):
    label     = JENIS_LABEL[jenis]
    ws        = wb.create_sheet(f"Detail-{label.replace('Simpanan ','')}")
    _no_grid(ws)
    # 9 kolom: No | No.Ang | Nama | Periode | Bulan | Tahun | Jumlah | Tgl Bayar | Keterangan
    NCOLS = 9
    _title_block(ws, "KOPERASI LANGGENG", f"DETAIL TRANSAKSI – {label.upper()}",
                 _periode_info(conn, periode_id, tahun, bulan), NCOLS)
    _header_row(ws,
        ["No","No. Anggota","Nama Anggota","Periode",
         "Bulan","Tahun","Jumlah (Rp)","Tgl Bayar","Keterangan"], 4)

    wh, wv = _where(periode_id, tahun, bulan)
    # tgl adalah NOT NULL di schema, jadi langsung ambil s.tgl
    q = f"""
        SELECT a.no_anggota, a.nama,
               COALESCE(pr.nama, '-')  AS periode_nama,
               s.bulan, s.tahun, s.jumlah,
               s.tgl,
               s.keterangan
        FROM simpanan s
        JOIN anggota a ON s.anggota_id = a.id
        LEFT JOIN periode pr ON s.periode_id = pr.id
        WHERE s.jenis = ?
        {'AND ' + wh[6:] if wh else ''}
        ORDER BY s.tgl ASC, a.nama ASC
    """
    rows = conn.execute(q, [jenis] + wv).fetchall()
    total = 0
    for i, r in enumerate(rows, 1):
        bln_txt = BULAN_NAMA.get(r[3], "-") if r[3] else "-"
        _data_cells(ws,
            [i, r[0], r[1], r[2], bln_txt, r[4] or "-",
             r[5], r[6] or "-", r[7] or "-"],
            i+4, num_cols=[7], alt=i%2==0)
        total += r[5]
    nr = len(rows)+5
    _total_cells(ws, ["","","TOTAL","","","", total,"",""], nr, num_cols=[7])
    _set_widths(ws, {"A":4,"B":11,"C":22,"D":18,"E":12,"F":7,"G":16,"H":13,"I":26})


# ═══════════════════════════════════════════════════════════════════════
#  SHEET: DETAIL SIMPANAN FORMAT MATRIKS (seperti gambar)
#  Kolom = tanggal bayar aktual dikelompokkan per bulan
#  Baris = per anggota, 1 periode saja
# ═══════════════════════════════════════════════════════════════════════
def _sheet_detail_simpanan_matriks(wb, conn, jenis, periode_id=None, tahun=None, bulan=None):
    """
    Format matriks seperti gambar referensi:
      Header baris 1-3: judul
      Baris 4: header gabung (No, No.Anggota, Nama, span Bulan per bulan, Jumlah, Ket)
      Baris 5: sub-header tanggal aktual per bulan
      Data: 1 baris per anggota, nilai di kolom tanggal bayar
    Hanya untuk 1 periode (periode_id / tahun / bulan tertentu).
    """
    label = JENIS_LABEL[jenis]
    ws    = wb.create_sheet(f"Detail-{label.replace('Simpanan ','')}")
    _no_grid(ws)

    # ── Ambil semua transaksi sesuai filter ──────────────────────────────
    wh, wv = _where(periode_id, tahun, bulan)
    q = f"""
        SELECT a.id AS anggota_id, a.no_anggota, a.nama,
               s.bulan, s.tahun, s.jumlah,
               CAST(strftime('%d', s.tgl) AS INTEGER) AS hari,
               s.tgl, s.keterangan
        FROM simpanan s
        JOIN anggota a ON s.anggota_id = a.id
        WHERE s.jenis = ?
        {'AND ' + wh[6:] if wh else ''}
        ORDER BY s.tgl ASC, a.nama ASC
    """
    rows = conn.execute(q, [jenis] + wv).fetchall()

    if not rows:
        # Buat sheet kosong dengan pesan
        NCOLS = 5
        _title_block(ws, "KOPERASI LANGGENG", f"DETAIL TRANSAKSI – {label.upper()}",
                     _periode_info(conn, periode_id, tahun, bulan), NCOLS)
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=NCOLS)
        c = ws.cell(row=5, column=1, value="Tidak ada data transaksi")
        c.font = _f(color=HEX["gray"], size=10); c.alignment = _al("center")
        _set_widths(ws, {"A":5,"B":13,"C":24,"D":16,"E":16})
        return

    # ── Kumpulkan bulan+tanggal unik ──────────────────────────────────────
    # bulan_tgl: dict {bulan: sorted list of hari unik}
    from collections import defaultdict, OrderedDict
    bulan_tgl = defaultdict(set)
    for r in rows:
        bln = r["bulan"] or 0
        hari = r["hari"] or 0
        if bln and hari:
            bulan_tgl[bln].add(hari)
    # Sort bulan, lalu hari dalam tiap bulan
    bulan_sorted = sorted(bulan_tgl.keys())
    bulan_tgl_sorted = OrderedDict()
    for bln in bulan_sorted:
        bulan_tgl_sorted[bln] = sorted(bulan_tgl[bln])

    # ── Hitung total kolom ────────────────────────────────────────────────
    # 3 kolom tetap: No | No.Anggota | Nama
    # lalu per bulan: N kolom tanggal
    # lalu 2 kolom tetap: Jumlah (Rp) | Keterangan
    FIXED_L = 3   # No, No.Anggota, Nama
    FIXED_R = 2   # Jumlah (Rp), Keterangan
    total_tgl_cols = sum(len(v) for v in bulan_tgl_sorted.values())
    NCOLS = FIXED_L + total_tgl_cols + FIXED_R

    # ── Judul ─────────────────────────────────────────────────────────────
    _title_block(ws, "KOPERASI LANGGENG", f"DETAIL TRANSAKSI – {label.upper()}",
                 _periode_info(conn, periode_id, tahun, bulan), NCOLS)

    # ── Baris 4: header grup bulan ────────────────────────────────────────
    # Kolom 1-3 merge row 4-5 (No, No.Anggota, Nama)
    static_hdrs = [("No", 1), ("No. Anggota", 2), ("Nama Anggota", 3)]
    for txt, col in static_hdrs:
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        c = ws.cell(row=4, column=col, value=txt)
        c.font = _f(bold=True, color="FFFFFF", size=9)
        c.fill = _fill(HEX["navy"]); c.alignment = _al("center")
        c.border = _bord_h(); ws.row_dimensions[4].height = 18

    # Span bulan di baris 4
    col_cursor = FIXED_L + 1
    bulan_col_start = {}  # bln -> col start
    alt_colors = [HEX["blue"], HEX["teal"]]
    for bi, (bln, hari_list) in enumerate(bulan_tgl_sorted.items()):
        n = len(hari_list)
        bcolor = alt_colors[bi % 2]
        bulan_col_start[bln] = col_cursor
        if n > 1:
            ws.merge_cells(start_row=4, start_column=col_cursor,
                           end_row=4, end_column=col_cursor + n - 1)
        c = ws.cell(row=4, column=col_cursor, value=BULAN_NAMA[bln])
        c.font = _f(bold=True, color="FFFFFF", size=9)
        c.fill = _fill(bcolor); c.alignment = _al("center"); c.border = _bord_h()
        col_cursor += n

    # Jumlah (Rp) dan Keterangan di baris 4 (merge row 4-5)
    col_jumlah = FIXED_L + total_tgl_cols + 1
    col_ket    = col_jumlah + 1
    for col, txt in [(col_jumlah, "Jumlah (Rp)"), (col_ket, "Keterangan")]:
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        c = ws.cell(row=4, column=col, value=txt)
        c.font = _f(bold=True, color="FFFFFF", size=9)
        c.fill = _fill(HEX["navy"]); c.alignment = _al("center")
        c.border = _bord_h()

    # ── Baris 5: sub-header tanggal ───────────────────────────────────────
    ws.row_dimensions[5].height = 16
    col_cursor = FIXED_L + 1
    for bi, (bln, hari_list) in enumerate(bulan_tgl_sorted.items()):
        bcolor = alt_colors[bi % 2]
        for hari in hari_list:
            c = ws.cell(row=5, column=col_cursor, value=hari)
            c.font = _f(bold=True, color="FFFFFF", size=9)
            c.fill = _fill(bcolor); c.alignment = _al("center"); c.border = _bord_h()
            col_cursor += 1

    # ── Lookup data: {(anggota_id, bulan, hari): jumlah, keterangan} ──────
    from collections import defaultdict
    # Satu anggota bisa bayar lebih dari 1x di tgl yang sama (edge case) → sum
    pay_data  = defaultdict(float)   # (aid, bln, hari) -> jumlah
    pay_ket   = {}                   # (aid, bln, hari) -> keterangan (ambil yg terakhir)
    anggota_info = {}                # aid -> (no_anggota, nama)
    aid_order = []
    for r in rows:
        aid  = r["anggota_id"]
        bln  = r["bulan"] or 0
        hari = r["hari"] or 0
        if aid not in anggota_info:
            anggota_info[aid] = (r["no_anggota"], r["nama"])
            aid_order.append(aid)
        if bln and hari:
            pay_data[(aid, bln, hari)] += r["jumlah"]
            pay_ket[(aid, bln, hari)]  = r["keterangan"] or ""

    # Keterangan per anggota = gabung semua unik
    ket_per_ang = defaultdict(set)
    for r in rows:
        if r["keterangan"]:
            ket_per_ang[r["anggota_id"]].add(r["keterangan"])

    # ── Data rows ────────────────────────────────────────────────────────
    # Buat map kolom: (bln, hari) -> col_index
    col_map = {}
    col_cursor = FIXED_L + 1
    for bln, hari_list in bulan_tgl_sorted.items():
        for hari in hari_list:
            col_map[(bln, hari)] = col_cursor
            col_cursor += 1

    num_cols_set = set(range(FIXED_L+1, col_jumlah+1))  # semua kolom angka

    grand_total = 0
    col_totals  = defaultdict(float)  # (bln,hari) -> total kolom
    for i, aid in enumerate(aid_order, 1):
        no_ang, nama = anggota_info[aid]
        row_excel = i + 5  # data mulai baris 6
        bg = HEX["alt"] if i % 2 == 0 else HEX["white"]

        # Kolom tetap kiri
        for col, val in [(1, i), (2, no_ang), (3, nama)]:
            c = ws.cell(row=row_excel, column=col, value=val)
            c.fill = _fill(bg); c.border = _bord(); c.font = _f(size=9)
            c.alignment = _al("right" if col == 1 else "left")

        # Kolom tanggal
        row_total = 0.0
        for (bln, hari), col in col_map.items():
            val = pay_data.get((aid, bln, hari), None)
            c = ws.cell(row=row_excel, column=col, value=val)
            c.fill = _fill(bg); c.border = _bord(); c.font = _f(size=9)
            c.alignment = _al("right")
            if val is not None:
                c.number_format = FMT_RP
                row_total += val
                col_totals[(bln, hari)] += val
            else:
                c.number_format = '"-"'

        # Jumlah total baris
        c_jml = ws.cell(row=row_excel, column=col_jumlah, value=row_total)
        c_jml.fill = _fill(bg); c_jml.border = _bord(); c_jml.font = _f(size=9, bold=True)
        c_jml.alignment = _al("right"); c_jml.number_format = FMT_RP
        grand_total += row_total

        # Keterangan
        ket_txt = ", ".join(sorted(ket_per_ang[aid])) if ket_per_ang[aid] else "-"
        c_ket = ws.cell(row=row_excel, column=col_ket, value=ket_txt)
        c_ket.fill = _fill(bg); c_ket.border = _bord(); c_ket.font = _f(size=9)
        c_ket.alignment = _al("left")

    # ── Baris total ───────────────────────────────────────────────────────
    total_row = len(aid_order) + 6
    for col in range(1, NCOLS + 1):
        c = ws.cell(row=total_row, column=col)
        c.font = _f(bold=True, color=HEX["total_fg"], size=9)
        c.fill = _fill(HEX["total_bg"]); c.border = _bord_h()
        c.alignment = _al("center")

    ws.cell(row=total_row, column=3, value="TOTAL")
    ws.cell(row=total_row, column=3).alignment = _al("center")

    for (bln, hari), col in col_map.items():
        val = col_totals.get((bln, hari), 0)
        c = ws.cell(row=total_row, column=col, value=val if val else None)
        c.font = _f(bold=True, color=HEX["total_fg"], size=9)
        c.fill = _fill(HEX["total_bg"]); c.border = _bord_h()
        c.alignment = _al("right")
        c.number_format = FMT_RP if val else '"-"'

    c_tot = ws.cell(row=total_row, column=col_jumlah, value=grand_total)
    c_tot.font = _f(bold=True, color=HEX["total_fg"], size=9)
    c_tot.fill = _fill(HEX["total_bg"]); c_tot.border = _bord_h()
    c_tot.alignment = _al("right"); c_tot.number_format = FMT_RP
    ws.row_dimensions[total_row].height = 18

    # ── Lebar kolom ───────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 24
    # Kolom tanggal: lebar tipis
    for col_idx in range(FIXED_L + 1, col_jumlah):
        ws.column_dimensions[_col(col_idx)].width = 11
    ws.column_dimensions[_col(col_jumlah)].width = 14
    ws.column_dimensions[_col(col_ket)].width = 18



# ═══════════════════════════════════════════════════════════════════════
#  SHEET: MATRIKS SIMPANAN (Jan-Des) per tahun
# ═══════════════════════════════════════════════════════════════════════
def _sheet_matriks_simpanan(wb, conn, jenis, tahun):
    label = JENIS_LABEL[jenis]
    ws    = wb.create_sheet(f"Matriks-{label.replace('Simpanan ','')}")
    _no_grid(ws)
    ncols = 16
    _title_block(ws, "KOPERASI LANGGENG",
                 f"MATRIKS {label.upper()} TAHUN {tahun}",
                 f"Tahun: {tahun}  |  Dicetak: {date.today().strftime('%d %B %Y')}", ncols)
    hdrs = ["No","No. Anggota","Nama Anggota"] + \
           [BULAN_SHORT[b] for b in range(1,13)] + ["TOTAL"]
    col_colors = [HEX["navy"]]*3 + \
                 [HEX["blue"] if b%2==1 else HEX["teal"] for b in range(1,13)] + \
                 [HEX["navy"]]
    _header_row(ws, hdrs, 4, col_colors)

    anggota = conn.execute("SELECT * FROM anggota ORDER BY id").fetchall()
    raw = conn.execute("""
        SELECT anggota_id, bulan, SUM(jumlah)
        FROM simpanan
        WHERE jenis=? AND tahun=?
        GROUP BY anggota_id, bulan
    """, (jenis, tahun)).fetchall()
    lookup = {(r[0], r[1]): r[2] for r in raw if r[1]}

    grand = 0; col_tot = {b: 0 for b in range(1,13)}
    for i, a in enumerate(anggota, 1):
        rv = [i, a["no_anggota"], a["nama"]]
        rt = 0
        for b in range(1,13):
            v = lookup.get((a["id"], b), None)
            rv.append(v); rt += (v or 0); col_tot[b] += (v or 0)
        rv.append(rt); grand += rt
        _data_cells(ws, rv, i+4, num_cols=list(range(4,17)), alt=i%2==0)
        for col in range(4, 16):
            c = ws.cell(row=i+4, column=col)
            if c.value is None: c.number_format = '"-"'
            else: c.number_format = FMT_RP_Z

    nr = len(anggota)+5
    _total_cells(ws,
        ["","","TOTAL"] + [col_tot[b] for b in range(1,13)] + [grand],
        nr, num_cols=list(range(4,17)))
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 22
    for b in range(1,13): ws.column_dimensions[_col(b+3)].width = 10
    ws.column_dimensions[_col(16)].width = 14


# ═══════════════════════════════════════════════════════════════════════
#  SHEET: REKAP SIMPANAN PER ANGGOTA PER PERIODE
#
#  Layout (sesuai gambar):
#   Baris 1-3 : judul & info tahun/cetak
#   Baris 4   : No | No.Anggota | Nama | [Simp. Pokok (merge n_per)] | ... | Total | Keterangan
#   Baris 5   : sub-header: Periode 1 (tgl tutup) | Periode 2 | ...  (per jenis)
#   Data      : 1 baris per anggota
#   Baris akhir: TOTAL
#
#  Tgl tutup periode = 31 Desember tahun periode (sesuai permintaan)
#  + Sheet "Smp-Semua" (akumulatif, tidak berubah)
# ═══════════════════════════════════════════════════════════════════════
def _sheet_rekap_tahunan(wb, conn, periode_id=None, tahun=None):
    """
    Sheet Rekap Tahunan Simpanan mengikuti filter user:
      - periode_id dipilih  -> hanya kolom periode itu, tanpa sheet Smp-Semua
      - tahun dipilih       -> hanya periode dalam tahun itu, tanpa sheet Smp-Semua
      - tanpa filter        -> semua periode + sheet Smp-Semua tetap muncul
    """
    from collections import defaultdict
    from calendar import monthrange
    from datetime import datetime as _dt

    ada_filter = (periode_id is not None) or (tahun is not None)

    anggota = conn.execute("SELECT * FROM anggota ORDER BY id").fetchall()

    def _to_ddmmyyyy(tgl_str, fallback_day, fallback_m, fallback_y):
        if tgl_str:
            try:
                d = _dt.strptime(tgl_str, "%Y-%m-%d")
                return f"{d.day:02d}/{d.month:02d}/{d.year}"
            except Exception:
                pass
        return f"{fallback_day:02d}/{fallback_m:02d}/{fallback_y}"

    # Ambil periode sesuai filter
    if periode_id is not None:
        periode_db = conn.execute(
            "SELECT id, nama, tahun, bulan_mulai, bulan_akhir, tgl_mulai, tgl_akhir "
            "FROM periode WHERE id=?", (periode_id,)
        ).fetchall()
    elif tahun is not None:
        periode_db = conn.execute(
            "SELECT id, nama, tahun, bulan_mulai, bulan_akhir, tgl_mulai, tgl_akhir "
            "FROM periode WHERE tahun=? ORDER BY id ASC", (tahun,)
        ).fetchall()
    else:
        periode_db = conn.execute(
            "SELECT id, nama, tahun, bulan_mulai, bulan_akhir, tgl_mulai, tgl_akhir "
            "FROM periode ORDER BY tahun ASC, id ASC"
        ).fetchall()

    # Simpanan tanpa periode hanya muncul saat tidak ada filter
    no_periode_count = 0
    if not ada_filter:
        no_periode_count = conn.execute(
            "SELECT COUNT(*) FROM simpanan WHERE periode_id IS NULL"
        ).fetchone()[0]

    # Bangun periode_list
    periode_list = []
    for p in periode_db:
        if periode_id is not None:
            cnt = conn.execute(
                "SELECT COUNT(*) FROM simpanan WHERE periode_id=?", (p["id"],)
            ).fetchone()[0]
        elif tahun is not None:
            cnt = conn.execute(
                "SELECT COUNT(*) FROM simpanan WHERE periode_id=? AND tahun=?",
                (p["id"], tahun)
            ).fetchone()[0]
        else:
            cnt = conn.execute(
                "SELECT COUNT(*) FROM simpanan WHERE periode_id=?", (p["id"],)
            ).fetchone()[0]

        if cnt > 0 or (periode_id is not None and periode_id == p["id"]):
            hari_akhir = monthrange(p["tahun"], p["bulan_akhir"])[1]
            tgl_m_fmt  = _to_ddmmyyyy(p["tgl_mulai"],   1,          p["bulan_mulai"], p["tahun"])
            tgl_a_fmt  = _to_ddmmyyyy(p["tgl_akhir"], hari_akhir,   p["bulan_akhir"], p["tahun"])
            periode_list.append({
                "id"       : p["id"],
                "label"    : p["nama"],
                "sublabel" : f"{tgl_m_fmt} s/d {tgl_a_fmt}",
                "tgl_mulai": tgl_m_fmt,
                "tgl_akhir": tgl_a_fmt,
            })

    if no_periode_count > 0:
        periode_list.append({
            "id"       : None,
            "label"    : "Tanpa Periode",
            "sublabel" : "-",
            "tgl_mulai": "-",
            "tgl_akhir": "-",
        })

    if not periode_list:
        periode_list = [{"id": None, "label": "Tanpa Periode",
                         "sublabel": "-", "tgl_mulai": "-", "tgl_akhir": "-"}]

    # ── Jenis simpanan ────────────────────────────────────────────────────
    JENIS_SUB  = ["pokok",        "wajib",        "sukarela",        "hariraya",    "khusus"]
    JENIS_FULL = ["Simpanan Pokok","Simpanan Wajib","Simpanan Sukarela","Simpanan Hari Raya","Simpanan Khusus"]

    n_per   = len(periode_list)   # jumlah periode
    n_jenis = len(JENIS_SUB)      # 5

    # Struktur kolom:
    #   col 1         : No
    #   col 2         : No. Anggota
    #   col 3         : Nama Anggota
    #   col 4 .. 4+n_per-1          : jenis[0] (Pokok) → sub: Periode 1..n
    #   col 4+n_per .. 4+2*n_per-1  : jenis[1] (Wajib)
    #   ...
    #   col 4+(n_jenis-1)*n_per .. col_total-1 : jenis[4] (Khusus)
    #   col_total     : Total
    #   col_ket       : Keterangan
    FIXED_L   = 3
    col_total = FIXED_L + n_jenis * n_per + 1
    col_ket   = col_total + 1
    NCOLS     = col_ket

    # Helper: kolom untuk (jenis_idx, periode_idx) → col number (1-based)
    def col_of(ji, pi):
        return FIXED_L + ji * n_per + pi + 1

    # ════════════════════════════════════════════════════
    # SHEET 1: Rekap Simpanan (sesuai filter)
    # ════════════════════════════════════════════════════
    ws = wb.create_sheet("Rekap Tahunan Simpanan")
    _no_grid(ws)
    # Judul info menyesuaikan filter
    if periode_id is not None and periode_list:
        p0 = periode_list[0]
        info = f"Periode: {p0['label']}  ({p0['sublabel']})  |  Dicetak: {date.today().strftime('%d %B %Y')}"
        subtitle = f"REKAP SIMPANAN – {p0['label'].upper()}"
    elif tahun is not None:
        info = f"Tahun: {tahun}  |  Dicetak: {date.today().strftime('%d %B %Y')}"
        subtitle = f"REKAP SIMPANAN TAHUN {tahun}"
    else:
        info = f"Semua Periode  |  Dicetak: {date.today().strftime('%d %B %Y')}"
        subtitle = "REKAP SIMPANAN (SEMUA PERIODE)"
    _title_block(ws, "KOPERASI LANGGENG", subtitle, info, NCOLS)

    # ── Baris 4 & 5: header ───────────────────────────────────────────────
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 48  # cukup untuk 2 baris: nama + tgl range

    # Kolom tetap kiri: merge row 4-5
    for col, txt in [(1,"No"), (2,"No. Anggota"), (3,"Nama Anggota")]:
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        c = ws.cell(row=4, column=col, value=txt)
        c.font = _f(bold=True, color="FFFFFF", size=9)
        c.fill = _fill(HEX["navy"]); c.alignment = _al("center"); c.border = _bord_h()

    # Kolom per jenis: baris 4 = nama jenis (merge n_per kolom)
    #                  baris 5 = label periode + tgl tutup
    alt_jenis_colors = [HEX["blue"], HEX["teal"], HEX["purple"], HEX["amber"], HEX["green"]]
    for ji, (jenis_key, jenis_full) in enumerate(zip(JENIS_SUB, JENIS_FULL)):
        jcolor = alt_jenis_colors[ji % len(alt_jenis_colors)]
        col_s  = col_of(ji, 0)
        col_e  = col_of(ji, n_per - 1)

        # Baris 4: nama jenis (merge seluruh periode)
        if n_per > 1:
            ws.merge_cells(start_row=4, start_column=col_s, end_row=4, end_column=col_e)
        c4 = ws.cell(row=4, column=col_s, value=jenis_full)
        c4.font = _f(bold=True, color="FFFFFF", size=9)
        c4.fill = _fill(jcolor); c4.alignment = _al("center"); c4.border = _bord_h()

        # Baris 5: satu sel per periode — "Nama Periode\nDD/MM/YYYY s/d DD/MM/YYYY"
        for pi, pinfo in enumerate(periode_list):
            col = col_of(ji, pi)
            sub_txt = f"{pinfo['label']}\n{pinfo['sublabel']}"
            c5 = ws.cell(row=5, column=col, value=sub_txt)
            c5.font = _f(bold=True, color="FFFFFF", size=8)
            c5.fill = _fill(jcolor); c5.alignment = _al("center", wrap=True); c5.border = _bord_h()

    # Total & Keterangan: merge row 4-5
    for col, txt in [(col_total, "Total"), (col_ket, "Keterangan")]:
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        c = ws.cell(row=4, column=col, value=txt)
        c.font = _f(bold=True, color="FFFFFF", size=9)
        c.fill = _fill(HEX["navy"]); c.alignment = _al("center", wrap=True); c.border = _bord_h()

    # ── Pre-fetch data simpanan ────────────────────────────────────────────
    raw = conn.execute("""
        SELECT anggota_id, periode_id, jenis, SUM(jumlah) as total
        FROM simpanan
        GROUP BY anggota_id, periode_id, jenis
    """).fetchall()
    lookup = defaultdict(float)
    for r in raw:
        lookup[(r["anggota_id"], r["periode_id"], r["jenis"])] = r["total"]

    # Keterangan per anggota (pinjaman aktif)
    pinjaman_aktif = {
        r[0] for r in conn.execute(
            "SELECT DISTINCT anggota_id FROM pinjaman WHERE status='aktif'"
        ).fetchall()
    }

    # ── Data rows ─────────────────────────────────────────────────────────
    col_totals_sum = defaultdict(float)

    for i, a in enumerate(anggota, 1):
        row_excel = i + 5
        bg = HEX["alt"] if i % 2 == 0 else HEX["white"]

        # Kolom tetap kiri
        for col, val in [(1, i), (2, a["no_anggota"]), (3, a["nama"])]:
            c = ws.cell(row=row_excel, column=col, value=val)
            c.fill = _fill(bg); c.border = _bord(); c.font = _f(size=9)
            c.alignment = _al("right" if col == 1 else "left")

        row_total = 0.0
        for ji, jenis_key in enumerate(JENIS_SUB):
            for pi, pinfo in enumerate(periode_list):
                col = col_of(ji, pi)
                val = lookup.get((a["id"], pinfo["id"], jenis_key), 0.0)
                c = ws.cell(row=row_excel, column=col, value=val if val else None)
                c.fill = _fill(bg); c.border = _bord(); c.font = _f(size=9)
                c.alignment = _al("right")
                if val:
                    c.number_format = FMT_RP
                    row_total += val
                    col_totals_sum[col] += val
                else:
                    c.number_format = '"-"'

        # Kolom Total baris
        c_t = ws.cell(row=row_excel, column=col_total, value=row_total if row_total else None)
        c_t.fill = _fill(bg); c_t.border = _bord()
        c_t.font = _f(size=9, bold=True); c_t.alignment = _al("right")
        c_t.number_format = FMT_RP if row_total else '"-"'
        if row_total:
            col_totals_sum[col_total] += row_total

        # Keterangan
        ket = "Ada Pinjaman" if a["id"] in pinjaman_aktif else ""
        c_k = ws.cell(row=row_excel, column=col_ket, value=ket)
        c_k.fill = _fill(bg); c_k.border = _bord()
        c_k.font = _f(size=9); c_k.alignment = _al("left")

    # ── Baris TOTAL ───────────────────────────────────────────────────────
    total_row = len(anggota) + 6
    ws.row_dimensions[total_row].height = 18
    for col in range(1, NCOLS + 1):
        c = ws.cell(row=total_row, column=col)
        c.font = _f(bold=True, color=HEX["total_fg"], size=9)
        c.fill = _fill(HEX["total_bg"]); c.border = _bord_h()
        c.alignment = _al("center")

    ws.cell(row=total_row, column=3, value="TOTAL")

    for col in range(FIXED_L + 1, col_ket):
        val = col_totals_sum.get(col, 0)
        c = ws.cell(row=total_row, column=col, value=val if val else None)
        c.font = _f(bold=True, color=HEX["total_fg"], size=9)
        c.fill = _fill(HEX["total_bg"]); c.border = _bord_h()
        c.alignment = _al("right")
        c.number_format = FMT_RP if val else '"-"'

    # ── Lebar kolom ───────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 24
    for col_idx in range(FIXED_L + 1, col_total):
        ws.column_dimensions[_col(col_idx)].width = 22  # cukup untuk "DD/MM/YYYY s/d DD/MM/YYYY"
    ws.column_dimensions[_col(col_total)].width = 16
    ws.column_dimensions[_col(col_ket)].width = 18


    # ════════════════════════════════════════════════════
    # SHEET 2: Smp-Semua — hanya muncul jika tanpa filter
    # ════════════════════════════════════════════════════
    if not ada_filter:
        ws3 = wb.create_sheet("Smp-Semua")
        _no_grid(ws3)
        info3 = f"Total Akumulatif Semua Periode  |  Dicetak: {date.today().strftime('%d %B %Y')}"
        _title_block(ws3, "KOPERASI LANGGENG",
                     "TOTAL SIMPANAN SEMUA PERIODE (AKUMULATIF)", info3, 8)
        _header_row(ws3,
            ["No","Nama Anggota","Total Pokok","Total Wajib","Total Sukarela",
             "Total Hari Raya","Total Khusus","GRAND TOTAL"], 4)

        grand3 = {j: 0 for j in JENIS_LIST}
        for i, a in enumerate(anggota, 1):
            rv = [i, a["nama"]]
            rt = 0
            for jenis in ["pokok","wajib","sukarela","hariraya","khusus"]:
                val = conn.execute(
                    "SELECT COALESCE(SUM(jumlah),0) FROM simpanan WHERE anggota_id=? AND jenis=?",
                    (a["id"], jenis)
                ).fetchone()[0]
                rv.append(val); rt += val; grand3[jenis] += val
            rv.append(rt)
            _data_cells(ws3, rv, i+4, num_cols=list(range(3,9)), alt=i%2==0)

        gt3 = sum(grand3.values())
        _total_cells(ws3,
            ["","TOTAL",
             grand3["pokok"],grand3["wajib"],grand3["sukarela"],
             grand3["hariraya"],grand3["khusus"],gt3],
            len(anggota)+5, num_cols=list(range(3,9)))
        _set_widths(ws3, {"A":4,"B":26,"C":14,"D":14,"E":14,"F":14,"G":14,"H":16})


# ═══════════════════════════════════════════════════════════════════════
#  SHEET: PINJAMAN
# ═══════════════════════════════════════════════════════════════════════
def _sheet_pinjaman(wb, conn, periode_id=None, tahun=None, bulan=None):
    ws = wb.create_sheet("Pinjaman")
    _no_grid(ws)
    _title_block(ws, "KOPERASI LANGGENG", "DATA PINJAMAN ANGGOTA",
                 _periode_info(conn, periode_id, tahun, bulan), 10)
    _header_row(ws,
        ["No","No. Anggota","Nama","Jumlah Pinjaman","Jangka (bln)",
         "Bunga/bln","Angsuran/bln","Tgl Pinjam","Keterangan","Status"], 4)
    wh, wv = _where(periode_id, tahun, bulan, "p")
    rows = conn.execute(f"""
        SELECT a.no_anggota, a.nama, p.jumlah, p.jangka, p.bunga,
               p.tgl, p.keterangan, p.status
        FROM pinjaman p JOIN anggota a ON p.anggota_id=a.id
        {wh} ORDER BY p.tgl DESC
    """, wv).fetchall()
    total_jml = 0; aktif = 0; lunas = 0
    for i, r in enumerate(rows, 1):
        ang = hitung_angsuran(r[2], r[3], r[4])
        _data_cells(ws,
            [i, r[0], r[1], r[2], r[3], r[4]/100, ang, r[5], r[6] or "-", r[7].upper()],
            i+4, num_cols=[4,7], alt=i%2==0)
        ws.cell(i+4, 6).number_format = "0.00%"
        ws.cell(i+4, 7).number_format = FMT_RP_Z
        st = ws.cell(i+4, 10)
        if r[7] == "lunas":
            st.font = _f(bold=True, color=HEX["green"]); lunas+=1
        else:
            st.font = _f(bold=True, color=HEX["orange"]); aktif+=1
        total_jml += r[2]
    _total_cells(ws,
        ["","","TOTAL", total_jml,"","","","",
         f"Aktif:{aktif} Lunas:{lunas}", ""],
        len(rows)+5, num_cols=[4])
    _set_widths(ws,{"A":4,"B":12,"C":24,"D":16,"E":10,"F":10,"G":14,"H":13,"I":22,"J":10})


# ═══════════════════════════════════════════════════════════════════════
#  SHEET: ANGSURAN — PERBAIKAN 2: tampilkan Bunga (Rp) dan Total Bayar
# ═══════════════════════════════════════════════════════════════════════
def _sheet_angsuran(wb, conn, periode_id=None, tahun=None, bulan=None):
    ws = wb.create_sheet("Angsuran")
    _no_grid(ws)
    _title_block(ws, "KOPERASI LANGGENG", "RIWAYAT PEMBAYARAN ANGSURAN",
                 _periode_info(conn, periode_id, tahun, bulan), 11)
    _header_row(ws,
        ["No","No. Anggota","Nama","Jumlah Pinjaman",
         "Angsuran ke","Bulan","Tahun",
         "Pokok Angsuran (Rp)","Bunga 1% (Rp)","Total Bayar (Rp)","Tgl Bayar"], 4)
    wh, wv = _where(periode_id, tahun, bulan, "ag")
    rows = conn.execute(f"""
        SELECT a.no_anggota, a.nama, p.jumlah, p.bunga, ag.ke,
               ag.bulan, ag.tahun, ag.jumlah, ag.tgl
        FROM angsuran ag
        JOIN pinjaman p ON ag.pinjaman_id=p.id
        JOIN anggota a  ON p.anggota_id=a.id
        {wh} ORDER BY ag.tgl ASC, ag.id ASC
    """, wv).fetchall()
    tp = 0; tb = 0; ttot = 0
    for i, r in enumerate(rows, 1):
        bln_txt  = BULAN_NAMA.get(r[5], "-") if r[5] else "-"
        bunga_pct   = r[3] if r[3] else 1.0
        nom_bunga   = round(r[2] * bunga_pct / 100)
        nom_pokok   = r[7] - nom_bunga
        if nom_pokok < 0: nom_pokok = r[7]; nom_bunga = 0
        tot_row = nom_pokok + nom_bunga
        _data_cells(ws,
            [i, r[0], r[1], r[2], r[4], bln_txt, r[6] or "-",
             nom_pokok, nom_bunga, tot_row, r[8] or "-"],
            i+4, num_cols=[4,8,9,10], alt=i%2==0)
        tp += nom_pokok; tb += nom_bunga; ttot += tot_row
    _total_cells(ws,
        ["","","TOTAL","", f"{len(rows)} kali","","", tp, tb, ttot, ""],
        len(rows)+5, num_cols=[4,8,9,10])
    _set_widths(ws,{"A":4,"B":12,"C":24,"D":16,"E":10,"F":12,"G":7,
                    "H":18,"I":14,"J":16,"K":13})


def _sheet_angsuran_matriks(wb, conn, tahun):
    ws = wb.create_sheet("Angsuran-Matriks")
    _no_grid(ws)
    ncols = 16
    _title_block(ws, "KOPERASI LANGGENG",
                 f"REKAPITULASI ANGSURAN TAHUN {tahun}",
                 f"Tahun: {tahun}  |  Dicetak: {date.today().strftime('%d %B %Y')}", ncols)
    hdrs = ["No","No. Anggota","Nama"] + \
           [BULAN_SHORT[b] for b in range(1,13)] + ["TOTAL"]
    cc = [HEX["navy"]]*3 + \
         [HEX["blue"] if b%2==1 else HEX["teal"] for b in range(1,13)] + [HEX["navy"]]
    _header_row(ws, hdrs, 4, cc)

    pinjaman = conn.execute("""
        SELECT p.id, a.no_anggota, a.nama, p.keterangan
        FROM pinjaman p JOIN anggota a ON p.anggota_id=a.id ORDER BY a.nama, p.id
    """).fetchall()
    raw = conn.execute("""
        SELECT pinjaman_id, bulan, SUM(jumlah)
        FROM angsuran WHERE tahun=? GROUP BY pinjaman_id, bulan
    """, (tahun,)).fetchall()
    lkp = {(r[0],r[1]): r[2] for r in raw if r[1]}

    grand = 0; ct = {b:0 for b in range(1,13)}
    for i, p in enumerate(pinjaman, 1):
        lbl = f"{p[2]} ({p[3] or 'Pin.'+str(p[0])})"
        rv = [i, p[1], lbl]; rt = 0
        for b in range(1,13):
            v = lkp.get((p[0],b), None)
            rv.append(v); rt += (v or 0); ct[b] += (v or 0)
        rv.append(rt); grand += rt
        _data_cells(ws, rv, i+4, num_cols=list(range(4,17)), alt=i%2==0)
        for col in range(4,16):
            c = ws.cell(row=i+4, column=col)
            if c.value is None: c.number_format = '"-"'
            else: c.number_format = FMT_RP_Z

    _total_cells(ws, ["","","TOTAL"]+[ct[b] for b in range(1,13)]+[grand],
                 len(pinjaman)+5, num_cols=list(range(4,17)))
    ws.column_dimensions["A"].width = 4; ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 28
    for b in range(1,13): ws.column_dimensions[_col(b+3)].width = 10
    ws.column_dimensions[_col(16)].width = 14


def _sheet_neraca(wb, conn, periode_id=None, tahun=None):
    ws = wb.create_sheet("Neraca")
    _no_grid(ws)
    _title_block(ws, "KOPERASI LANGGENG", "NERACA KEUANGAN (RINGKASAN)",
                 _periode_info(conn, periode_id, tahun), 5)
    ws.row_dimensions[5].height = 6

    def sec(row, label, bg=HEX["navy"]):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=label)
        c.font=_f(bold=True, color="FFFFFF", size=10); c.fill=_fill(bg)
        c.alignment=_al("left"); ws.row_dimensions[row].height=18

    def item(row, label, value, bold=False, bg=HEX["white"], indent=False):
        lbl = ("      " if indent else "") + label
        for ci in range(1, 6):
            c = ws.cell(row=row, column=ci); c.fill=_fill(bg); c.border=_bord()
        c1 = ws.cell(row=row, column=1, value=lbl)
        c1.font=_f(bold=bold, size=9); c1.alignment=_al("left")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cv = ws.cell(row=row, column=5, value=value)
        cv.font=_f(bold=bold, size=9); cv.number_format=FMT_RP_Z; cv.alignment=_al("right")

    wh, wv   = _where(periode_id, tahun, None, "s")
    kas      = conn.execute(f"SELECT SUM(jumlah) FROM simpanan s {wh}", wv).fetchone()[0] or 0
    wh2, wv2 = _where(periode_id, tahun, None, "p")
    awh = (wh2 + " AND p.status='aktif'") if wh2 else "WHERE p.status='aktif'"
    pin_aktif = conn.execute(f"SELECT SUM(jumlah) FROM pinjaman p {awh}", wv2).fetchone()[0] or 0
    ang_masuk = conn.execute(
        f"SELECT SUM(ag.jumlah) FROM angsuran ag "
        f"JOIN pinjaman p ON ag.pinjaman_id=p.id {wh2}", wv2).fetchone()[0] or 0
    piutang = max(0, pin_aktif - ang_masuk)

    r = 6
    sec(r, "  A S E T"); r+=1
    item(r, "Kas & Dana Simpanan", kas, indent=True); r+=1
    item(r, "Piutang Pinjaman (sisa)", piutang, indent=True); r+=1
    item(r, "TOTAL ASET", kas+piutang, bold=True, bg=HEX["blue_mid"]); r+=2
    sec(r, "  KEWAJIBAN", bg=HEX["teal"]); r+=1
    item(r, "Simpanan Anggota", kas, indent=True); r+=1
    item(r, "TOTAL KEWAJIBAN", kas, bold=True, bg=HEX["blue_lt"]); r+=2
    sec(r, "  MODAL / EKUITAS", bg=HEX["green"]); r+=1
    item(r, "Modal Beredar (Piutang Aktif)", piutang, indent=True); r+=1
    item(r, "TOTAL MODAL", piutang, bold=True, bg=HEX["green_lt"]); r+=2
    sec(r, "  RINGKASAN", bg=HEX["amber"]); r+=1
    total_pin = conn.execute(f"SELECT SUM(jumlah) FROM pinjaman p {wh2}", wv2).fetchone()[0] or 0
    jml_ang   = conn.execute("SELECT COUNT(*) FROM anggota").fetchone()[0]
    lwh = (wh2 + " AND p.status='lunas'") if wh2 else "WHERE p.status='lunas'"
    awh2= (wh2 + " AND p.status='aktif'") if wh2 else "WHERE p.status='aktif'"
    item(r, "Jumlah Anggota", jml_ang, indent=True); r+=1
    item(r, "Total Pinjaman Disalurkan", total_pin, indent=True); r+=1
    item(r, "Pinjaman Aktif",
         conn.execute(f"SELECT COUNT(*) FROM pinjaman p {awh2}", wv2).fetchone()[0],
         indent=True); r+=1
    item(r, "Pinjaman Lunas",
         conn.execute(f"SELECT COUNT(*) FROM pinjaman p {lwh}", wv2).fetchone()[0],
         indent=True); r+=1
    item(r, "Total Angsuran Diterima", ang_masuk, indent=True); r+=1
    _set_widths(ws, {"A":34,"B":10,"C":10,"D":10,"E":18})

# ═══════════════════════════════════════════════════════════════════════
#  FUNGSI PUBLIK
# ═══════════════════════════════════════════════════════════════════════

def export_rekap_lengkap(conn, filepath, periode_id=None, tahun=None, bulan=None):
    from helpers import this_year
    t  = tahun or this_year()
    wb = Workbook(); wb.remove(wb.active)
    _sheet_anggota(wb, conn, periode_id)
    _sheet_rekap_simpanan(wb, conn, periode_id, tahun, bulan)
    for j in JENIS_LIST:
        _sheet_detail_simpanan_matriks(wb, conn, j, periode_id, tahun, bulan)
    for j in JENIS_LIST:
        _sheet_matriks_simpanan(wb, conn, j, t)
    _sheet_rekap_tahunan(wb, conn, periode_id=periode_id, tahun=tahun)
    _sheet_pinjaman(wb, conn, periode_id, tahun, bulan)
    _sheet_angsuran(wb, conn, periode_id, tahun, bulan)
    _sheet_angsuran_matriks(wb, conn, t)
    _sheet_neraca(wb, conn, periode_id, tahun)
    wb.save(filepath); return filepath

def export_simpanan(conn, filepath, jenis=None, periode_id=None, tahun=None, bulan=None):
    from helpers import this_year
    t  = tahun or this_year()
    wb = Workbook(); wb.remove(wb.active)
    _sheet_rekap_simpanan(wb, conn, periode_id, tahun, bulan)
    target = [jenis] if jenis else JENIS_LIST
    for j in target:
        _sheet_detail_simpanan_matriks(wb, conn, j, periode_id, tahun, bulan)
    for j in target:
        _sheet_matriks_simpanan(wb, conn, j, t)
    _sheet_rekap_tahunan(wb, conn, periode_id=periode_id, tahun=tahun)
    wb.save(filepath); return filepath

def export_simpanan_perbulan(conn, filepath, jenis=None, tahun=None):
    from helpers import this_year
    t  = tahun or this_year()
    wb = Workbook(); wb.remove(wb.active)
    target = [jenis] if jenis else JENIS_LIST
    for j in target:
        _sheet_matriks_simpanan(wb, conn, j, t)
    _sheet_rekap_tahunan(wb, conn, tahun=tahun)
    wb.save(filepath); return filepath

def export_pinjaman_angsuran(conn, filepath, periode_id=None, tahun=None, bulan=None):
    from helpers import this_year
    t  = tahun or this_year()
    wb = Workbook(); wb.remove(wb.active)
    _sheet_pinjaman(wb, conn, periode_id, tahun, bulan)
    _sheet_angsuran(wb, conn, periode_id, tahun, bulan)
    _sheet_angsuran_matriks(wb, conn, t)
    wb.save(filepath); return filepath

# Alias agar import lama tidak error
export_pinjaman = export_pinjaman_angsuran
export_angsuran = export_pinjaman_angsuran

def export_anggota(conn, filepath):
    wb = Workbook(); wb.remove(wb.active)
    _sheet_anggota(wb, conn)
    wb.save(filepath); return filepath

def export_neraca(conn, filepath, periode_id=None, tahun=None):
    wb = Workbook(); wb.remove(wb.active)
    _sheet_neraca(wb, conn, periode_id, tahun)
    wb.save(filepath); return filepath



# ═══════════════════════════════════════════════════════════════════════
#  SHEET: KARTU ANGSURAN — format PERSIS seperti referensi gambar
#
#  Layout kolom:
#    A=No  B=Tgl.pinjam  C=Jumlah pinjaman  D=Jumlah Angsuran(label)
#    E=label(Pokok/jasa/Dibayar/ags/tgl)
#    F..Q = Jan..Des   R=Jumlah Bunga   S=Saldo
#
#  Baris 1 di atas kolom bulan: merge "TAHUN – nama anggota"
#
#  Per pinjaman = 5 baris AKTIF + 2 baris HISTORIS MERAH (opsional):
#    baris1 Pokok   : A=no(merge5) B=tgl(merge2) C=jumlah(merge5) D=(kosong) E=Pokok   | bulan=pokok/bln
#    baris2 Jasa    :              B=Rp.jml total               D=(kosong) E=jasa    | bulan=bunga/bln
#    baris3 Dibayar :                                           D=jangka(Nx)E=Dibayar| bulan=total dibayar
#    baris4 Ags     : A=no(merge)  B=nama(merge2)               D=(kosong) E=ags     | bulan="jangka/ke"
#    baris5 Tgl     :              B=nama                        D=(kosong) E=tgl     | bulan=tgl bayar
#    -- jika ada historis tahun sebelumnya: --
#    baris6 MERAH   :              B=tgl akhir historis (merah)
#    baris7 MERAH   :              B=saldo historis    (merah)
#
#  "36/6" artinya: angsuran ke-36 terjadi di bulan Juni → jangka/bulan_ke
#  Angsuran fleksibel: jangka diambil dari DB (bisa 6,12,24,36,dst)
# ═══════════════════════════════════════════════════════════════════════
def _sheet_kartu_pinjaman(wb, conn, tahun_filter=None):
    from datetime import datetime as _dt

    thn  = tahun_filter or date.today().year
    BULAN_NAMA_LIST = ["Januari","Feb","Maret","April","Mei","Juni",
                       "Juli","Agustus","September","oktober","Nov","Desember"]

    # ── Kolom ────────────────────────────────────────────────────────────
    C_NO  = 1   # A
    C_TGL = 2   # B
    C_JML = 3   # C
    C_AGS = 4   # D  (label "Jumlah Angsuran" di header; isi: kosong/jangka)
    C_LBL = 5   # E  (Pokok/jasa/Dibayar/ags/tgl)
    C_M1  = 6   # F  Jan
    C_M12 = 17  # Q  Des
    C_BNGA= 18  # R
    C_SLD = 19  # S
    NCOLS = 19

    # ── Warna ────────────────────────────────────────────────────────────
    NAVY    = "1E3A5F"
    BLUE    = "2B6CB0"
    LBLUE   = "2C7DBE"
    HDR_BG  = "D6E4F0"
    ALT_BG  = "EBF4FB"
    WHT     = "FFFFFF"
    HIST_BG = "FFF0F0"
    HIST_FG = "CC0000"
    GRN     = "1A6B3A"
    ORG     = "B54708"

    def _s(style="thin", color="B0C4DE"):
        return Side(style=style, color=color)

    def _brd(top="thin",bot="thin",lft="thin",rgt="thin",
             tc="B0C4DE",bc="B0C4DE",lc="B0C4DE",rc="B0C4DE"):
        return Border(top=_s(top,tc), bottom=_s(bot,bc),
                      left=_s(lft,lc), right=_s(rgt,rc))

    B_NORM = _brd()
    B_BOT  = _brd(bot="medium", bc=NAVY)

    def _wr(ws, row, col, val="", bold=False, fg="000000", bg=WHT,
            ha="center", sz=9, fmt=None, wrap=False, brd=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=bold, color=fg, size=sz, name="Arial")
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
        c.border    = brd if brd else B_NORM
        if fmt: c.number_format = fmt
        return c

    def _mg(ws, r1,c1,r2,c2, val="", bold=False, fg="000000", bg=WHT,
            ha="center", sz=9, fmt=None, brd=None):
        try:
            ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
        except Exception:
            pass
        c = ws.cell(row=r1, column=c1, value=val)
        c.font      = Font(bold=bold, color=fg, size=sz, name="Arial")
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=ha, vertical="center")
        c.border    = brd if brd else B_NORM
        if fmt: c.number_format = fmt
        return c

    def _fmt_tgl(s, fmt="%d/%m/%Y"):
        try:    return _dt.strptime(str(s), "%Y-%m-%d").strftime(fmt)
        except: return str(s) if s else ""

    def _rp(v):
        try:    return f"Rp {int(v):,.0f}".replace(",",".")
        except: return ""

    # Ambil anggota yang punya pinjaman
    anggota_list = conn.execute("""
        SELECT DISTINCT a.id, a.no_anggota, a.nama
        FROM anggota a JOIN pinjaman p ON p.anggota_id=a.id
        ORDER BY a.no_anggota
    """).fetchall()

    for ang in anggota_list:
        aid, no_ang, nama_ang = ang["id"], ang["no_anggota"], ang["nama"]

        pinjaman_list = conn.execute("""
            SELECT id,jumlah,jangka,bunga,tgl,status,keterangan
            FROM pinjaman WHERE anggota_id=?
            ORDER BY tgl ASC, id ASC
        """, (aid,)).fetchall()
        if not pinjaman_list: continue

        # Nama sheet = nama anggota
        sname = nama_ang[:31]
        existing = [w.title for w in wb.worksheets]
        sfx = 1
        while sname in existing:
            sname = f"{nama_ang[:28]}_{sfx}"; sfx += 1
        ws = wb.create_sheet(sname)
        _no_grid(ws)

        # ════ Baris 1: Header nama koperasi ══════════════════════════════
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
        c = ws.cell(row=1, column=1,
                    value=f"KARTU ANGSURAN PINJAMAN – {nama_ang.upper()}")
        c.font = Font(bold=True, color=WHT, size=12, name="Arial")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # ════ Baris 2: Sub-header info ════════════════════════════════════
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
        c = ws.cell(row=2, column=1,
                    value=f"No. Anggota: {no_ang}  |  Tahun: {thn}  |  "
                          f"Dicetak: {date.today().strftime('%d %B %Y')}")
        c.font = Font(bold=False, color=NAVY, size=9, name="Arial")
        c.fill = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 15

        # ════ Baris 3: Header kolom baris-1 (merge atas bulan) ════════════
        # Col A-E kosong di baris 3; Col F-Q merge "TAHUN – nama user"
        for ci in range(1, 6):
            _wr(ws, 3, ci, "", bg=NAVY, brd=B_NORM)
        _mg(ws, 3, C_M1, 3, C_M12,
            val=f"{thn}  –  {nama_ang}",
            bold=True, fg=WHT, bg=BLUE, ha="center", sz=10)
        _wr(ws, 3, C_BNGA, "", bg=NAVY, brd=B_NORM)
        _wr(ws, 3, C_SLD,  "", bg=NAVY, brd=B_NORM)
        ws.row_dimensions[3].height = 18

        # ════ Baris 4: Header kolom baris-2 ══════════════════════════════
        HDR = 4
        hdr_vals = ["No","Tgl.pinjam","Jumlah pinjaman","Jumlah Angsuran",""]
        hdr_vals += BULAN_NAMA_LIST
        hdr_vals += ["Jumlah Bunga","Saldo"]
        hdr_bgs  = [NAVY]*5 + [BLUE if i%2==0 else LBLUE for i in range(12)] + [NAVY,NAVY]
        for ci, (v, bg) in enumerate(zip(hdr_vals, hdr_bgs), 1):
            c = ws.cell(row=HDR, column=ci, value=v)
            c.font      = Font(bold=True, color=WHT, size=9, name="Arial")
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = B_NORM
        ws.row_dimensions[HDR].height = 28

        # ════ Baris data ══════════════════════════════════════════════════
        cur = HDR + 1

        for no_urut, p in enumerate(pinjaman_list, 1):
            pid       = p["id"]
            jumlah    = p["jumlah"]
            jangka    = p["jangka"]   # fleksibel: 6,12,24,36,dst
            bpct      = p["bunga"] or 1.5
            tgl_pin   = p["tgl"]
            status    = p["status"]
            ket       = p["keterangan"] or ""

            pokok_nom = round(jumlah / jangka) if jangka > 0 else 0
            bunga_nom = round(jumlah * bpct / 100)
            total_nom = pokok_nom + bunga_nom

            # Ambil semua angsuran
            angsur_all = conn.execute("""
                SELECT ke,bulan,tahun,jumlah,tgl FROM angsuran
                WHERE pinjaman_id=? ORDER BY ke ASC
            """, (pid,)).fetchall()

            # Kelompok: curr=tahun ini, hist=tahun sebelumnya
            # key=bulan, value=list angsuran (bisa >1 di bulan yg sama)
            curr = {}   # bulan → {ke,jumlah,tgl}  (ambil pertama)
            hist_list = []  # semua angsuran tahun < thn

            total_bunga_bayar = 0
            total_pokok_bayar = 0

            for ag in angsur_all:
                bln, thn_ag = ag["bulan"], ag["tahun"]
                if not bln: continue
                d = {"ke": ag["ke"], "jumlah": ag["jumlah"], "tgl": ag["tgl"] or ""}
                total_bunga_bayar += bunga_nom
                total_pokok_bayar += pokok_nom
                if thn_ag == thn:
                    if bln not in curr:
                        curr[bln] = d
                elif thn_ag and thn_ag < thn:
                    hist_list.append(ag)

            saldo = max(0.0, jumlah - total_pokok_bayar)
            if status == "lunas": saldo = 0.0

            # Hitung total bunga di tahun ini saja
            bunga_thn_ini = bunga_nom * len(curr)

            # Info historis: tgl akhir bayar & total dibayar di tahun sebelumnya
            hist_tgl_akhir  = ""
            hist_total_bayar= 0
            if hist_list:
                hist_sorted = sorted(hist_list, key=lambda x: (x["tahun"] or 0, x["ke"] or 0))
                last_h      = hist_sorted[-1]
                hist_tgl_akhir  = _fmt_tgl(last_h["tgl"], "%d %b %Y").upper()
                hist_total_bayar= sum((h["jumlah"] or 0) for h in hist_sorted)

            ada_hist = bool(hist_list)
            NROWS = 5 + (2 if ada_hist else 0)

            # Row indeks
            R1 = cur        # Pokok
            R2 = cur+1      # jasa
            R3 = cur+2      # Dibayar
            R4 = cur+3      # ags
            R5 = cur+4      # tgl
            R6 = cur+5      # historis baris1 (merah) — hanya jika ada hist
            R7 = cur+6      # historis baris2 (merah) — hanya jika ada hist

            for ri in range(cur, cur+NROWS):
                ws.row_dimensions[ri].height = 15
            if ada_hist:
                ws.row_dimensions[R6].height = 14
                ws.row_dimensions[R7].height = 14

            last_r = cur + NROWS - 1  # baris terakhir blok ini

            bg_norm = ALT_BG if no_urut % 2 == 0 else WHT

            # ── Kolom A: No (merge R1..R5) ────────────────────────────────
            _mg(ws, R1, C_NO, R5, C_NO,
                val=no_urut, bold=True, fg=BLUE, bg=bg_norm,
                ha="center", sz=12,
                brd=_brd(bot="hair",bc="B0C4DE") if ada_hist else B_BOT)
            if ada_hist:
                _mg(ws, R6, C_NO, R7, C_NO, val="", bg=HIST_BG, brd=B_BOT)

            # ── Kolom B: Tgl pinjam (baris1-2 merge) + nama user ─────────
            # B baris1-2: tgl pinjam
            _mg(ws, R1, C_TGL, R2, C_TGL,
                val=_fmt_tgl(tgl_pin, "%d %b %Y"), bold=False,
                fg="374151", bg=bg_norm, ha="center", sz=9)
            # B baris3: Rp. jumlah total bayar (jumlah + bunga)
            total_semua = jumlah + (bunga_nom * jangka)
            _wr(ws, R3, C_TGL,
                f"Rp. {int(jumlah):,}".replace(",","."),
                bold=False, fg="374151", bg=bg_norm, ha="left", sz=9)
            # B baris4-5: nama user (merge)
            _mg(ws, R4, C_TGL, R5, C_TGL,
                val=nama_ang, bold=False,
                fg="374151", bg=bg_norm, ha="center", sz=9)
            # B baris6-7 historis (merah)
            if ada_hist:
                _wr(ws, R6, C_TGL, hist_tgl_akhir,
                    bold=False, fg=HIST_FG, bg=HIST_BG, ha="center", sz=9)
                _wr(ws, R7, C_TGL,
                    f"Rp. {int(hist_total_bayar):,}".replace(",","."),
                    bold=True, fg=HIST_FG, bg=HIST_BG, ha="left", sz=9,
                    brd=B_BOT)

            # ── Kolom C: Jumlah pinjaman (merge R1..R5) ───────────────────
            _mg(ws, R1, C_JML, R5, C_JML,
                val=jumlah, bold=True, fg=NAVY, bg=bg_norm,
                ha="right", sz=9, fmt=FMT_RP,
                brd=_brd(bot="hair",bc="B0C4DE") if ada_hist else B_BOT)
            if ada_hist:
                _mg(ws, R6, C_JML, R7, C_JML, val="", bg=HIST_BG, brd=B_BOT)

            # ── Kolom D: Jangka & label ───────────────────────────────────
            # D baris3 saja: "36x" (jangka)
            for ri in [R1, R2, R4, R5]:
                _wr(ws, ri, C_AGS, "", bg=bg_norm)
            _wr(ws, R3, C_AGS, f"{jangka}x",
                bold=True, fg=NAVY, bg=bg_norm, ha="center", sz=10)
            if ada_hist:
                _wr(ws, R6, C_AGS, "", bg=HIST_BG)
                _wr(ws, R7, C_AGS, "", bg=HIST_BG, brd=B_BOT)

            # ── Kolom E: Label ────────────────────────────────────────────
            lbl_rows_def = [
                (R1, "Pokok",   False, "374151"),
                (R2, "jasa",    False, "374151"),
                (R3, "Dibayar", True,  "374151"),
                (R4, "ags",     False, "374151"),
                (R5, "tgl",     False,
                     GRN if status=="lunas" else ORG),
            ]
            for ri, lv, lb, lf in lbl_rows_def:
                brd = B_BOT if (ri == R5 and not ada_hist) else B_NORM
                _wr(ws, ri, C_LBL, lv, bold=lb, fg=lf,
                    bg=bg_norm, ha="center", sz=9, brd=brd)
            if ada_hist:
                _wr(ws, R6, C_LBL, "", bg=HIST_BG)
                _wr(ws, R7, C_LBL, ket or status.upper(),
                    bold=True, fg=HIST_FG, bg=HIST_BG,
                    ha="center", sz=8, brd=B_BOT)

            # ── Kolom F..Q: bulan ─────────────────────────────────────────
            for bi, bln in enumerate(range(1, 13)):
                col_b = C_M1 + bi

                if bln in curr:
                    ag  = curr[bln]
                    cbg = ALT_BG if no_urut%2==0 else WHT
                    cfg = "1A202C"
                    brd5 = B_BOT if not ada_hist else B_NORM

                    # baris1 Pokok
                    c = _wr(ws, R1, col_b, _rp(pokok_nom),
                            fg=cfg, bg=cbg, ha="right", sz=9)
                    # baris2 jasa
                    c = _wr(ws, R2, col_b, _rp(bunga_nom),
                            fg=cfg, bg=cbg, ha="right", sz=9)
                    # baris3 Dibayar
                    c = _wr(ws, R3, col_b, _rp(total_nom),
                            bold=True, fg=cfg, bg=cbg, ha="right", sz=9)
                    # baris4 ags ke → "jangka/ke_bulan_ini"
                    # "36/6" = jangka 36, angsuran ke-36 terjadi di bulan 6
                    _wr(ws, R4, col_b, f"{jangka}/{ag['ke']}",
                        fg=cfg, bg=cbg, ha="center", sz=9)
                    # baris5 tgl
                    _wr(ws, R5, col_b, _fmt_tgl(ag["tgl"], "%d-%b-%y"),
                        fg=cfg, bg=cbg, ha="center", sz=8, brd=brd5)
                    if ada_hist:
                        _wr(ws, R6, col_b, "", bg=HIST_BG)
                        _wr(ws, R7, col_b, "", bg=HIST_BG, brd=B_BOT)
                else:
                    # Kosong
                    brd5 = B_BOT if not ada_hist else B_NORM
                    _wr(ws, R1, col_b, "", bg=bg_norm)
                    _wr(ws, R2, col_b, "", bg=bg_norm)
                    _wr(ws, R3, col_b, "", bg=bg_norm)
                    _wr(ws, R4, col_b, "", bg=bg_norm)
                    _wr(ws, R5, col_b, "", bg=bg_norm, brd=brd5)
                    if ada_hist:
                        _wr(ws, R6, col_b, "", bg=HIST_BG)
                        _wr(ws, R7, col_b, "", bg=HIST_BG, brd=B_BOT)

            # ── Kolom R: Jumlah Bunga (merge R1..R5) ──────────────────────
            _mg(ws, R1, C_BNGA, R5, C_BNGA,
                val=bunga_thn_ini if bunga_thn_ini else None,
                bold=True, fg=ORG if bunga_thn_ini else "999999",
                bg=bg_norm, ha="right", sz=9,
                fmt=FMT_RP if bunga_thn_ini else None,
                brd=_brd(bot="hair",bc="B0C4DE") if ada_hist else B_BOT)
            if ada_hist:
                _mg(ws, R6, C_BNGA, R7, C_BNGA, val="", bg=HIST_BG, brd=B_BOT)

            # ── Kolom S: Saldo (merge R1..R5) ─────────────────────────────
            _mg(ws, R1, C_SLD, R5, C_SLD,
                val=saldo, bold=True,
                fg=GRN if status=="lunas" else ORG,
                bg=bg_norm, ha="right", sz=9, fmt=FMT_RP,
                brd=_brd(bot="hair",bc="B0C4DE") if ada_hist else B_BOT)
            if ada_hist:
                _mg(ws, R6, C_SLD, R7, C_SLD, val="", bg=HIST_BG, brd=B_BOT)

            cur += NROWS

        # ── Lebar kolom ───────────────────────────────────────────────────
        ws.column_dimensions[_col(C_NO) ].width = 4
        ws.column_dimensions[_col(C_TGL)].width = 14
        ws.column_dimensions[_col(C_JML)].width = 14
        ws.column_dimensions[_col(C_AGS)].width = 9
        ws.column_dimensions[_col(C_LBL)].width = 9
        for bi in range(12):
            ws.column_dimensions[_col(C_M1+bi)].width = 11
        ws.column_dimensions[_col(C_BNGA)].width = 13
        ws.column_dimensions[_col(C_SLD) ].width = 13


def export_kartu_pinjaman(conn, filepath, tahun=None):
    """Export kartu angsuran pinjaman per anggota (satu sheet per anggota)."""
    from helpers import this_year
    thn = tahun or this_year()
    wb = Workbook(); wb.remove(wb.active)
    _sheet_kartu_pinjaman(wb, conn, tahun_filter=thn)
    wb.save(filepath)
    return filepath